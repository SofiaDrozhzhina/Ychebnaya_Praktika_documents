# api.py
import io
import textwrap
from flask import Blueprint, request, jsonify, send_file, current_app
from sqlalchemy import asc, desc, or_
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from docx import Document
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.utils import ImageReader
from models import SessionLocal,engine,Student, Course, Record
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload
from reportlab.platypus import Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_JUSTIFY

# --- Blueprints ---
api = Blueprint('api', __name__, url_prefix='/api')
documents_bp = Blueprint("documents", __name__, url_prefix="/documents")
excel_bp = Blueprint('excel_bp', __name__, url_prefix='/excel')
pdf_bp = Blueprint("pdf", __name__, url_prefix="/pdf")


# --- Вспомогательная функция сортировки ---
def parse_sort_order(param):
    if param == 'asc':
        return asc
    if param == 'desc':
        return desc
    return None


# ===============================
# === STUDENTS (Студенты) ======
# ===============================
@api.route('/students', methods=['GET'])
def list_students():
    """Получить список студентов с фильтрацией и сортировкой"""
    q = request.args.get('q', '').strip()
    sort = request.args.get('sort', 'default')
    db = SessionLocal()
    query = db.query(Student)
    if q:
        query = query.filter(Student.fio.ilike(f'%{q}%'))
    if sort == 'asc':
        query = query.order_by(asc(Student.date_of_birth))
    elif sort == 'desc':
        query = query.order_by(desc(Student.date_of_birth))
    students = [s.to_dict() for s in query.all()]
    db.close()
    return jsonify(students)


@api.route('/students', methods=['POST'])
def create_student():
    """Создать нового студента"""
    data = request.json
    db = SessionLocal()
    try:
        dob = datetime.fromisoformat(data['date_of_birth']).date()
    except Exception:
        return jsonify({"error": "Invalid date_of_birth"}), 400
    student = Student(fio=data['fio'], date_of_birth=dob, phone=data.get('phone'))
    db.add(student)
    db.commit()
    db.refresh(student)
    res = student.to_dict()
    db.close()
    return jsonify(res), 201


@api.route('/students/<int:student_id>', methods=['PUT'])
def update_student(student_id):
    """Обновить данные студента"""
    data = request.json
    db = SessionLocal()
    student = db.query(Student).get(student_id)
    if not student:
        db.close()
        return jsonify({"error": "Not found"}), 404
    student.fio = data.get('fio', student.fio)
    if data.get('date_of_birth'):
        student.date_of_birth = datetime.fromisoformat(data['date_of_birth']).date()
    student.phone = data.get('phone', student.phone)
    db.commit()
    res = student.to_dict()
    db.close()
    return jsonify(res)


@api.route('/students/<int:student_id>', methods=['DELETE'])
def delete_student(student_id):
    """Удалить студента"""
    db = SessionLocal()
    student = db.query(Student).get(student_id)
    if not student:
        db.close()
        return jsonify({"error": "Not found"}), 404
    db.delete(student)
    db.commit()
    db.close()
    return jsonify({"status": "deleted"})


# ===============================
# === COURSES (Курсы) ==========
# ===============================
@api.route('/courses', methods=['GET'])
def list_courses():
    """Список курсов"""
    q = request.args.get('q', '').strip()
    teacher = request.args.get('teacher', '').strip()
    db = SessionLocal()
    query = db.query(Course)
    if q:
        query = query.filter(Course.name.ilike(f'%{q}%'))
    if teacher:
        query = query.filter(Course.teacher == teacher)
    courses = [c.to_dict() for c in query.all()]
    db.close()
    return jsonify(courses)


@api.route('/courses', methods=['POST'])
def create_course():
    """Создать курс"""
    data = request.json
    db = SessionLocal()
    course = Course(name=data['name'], description=data.get('description'), teacher=data.get('teacher'))
    db.add(course)
    db.commit()
    db.refresh(course)
    res = course.to_dict()
    db.close()
    return jsonify(res), 201


@api.route('/courses/<int:course_id>', methods=['PUT'])
def update_course(course_id):
    """Обновить курс"""
    data = request.json
    db = SessionLocal()
    course = db.query(Course).get(course_id)
    if not course:
        db.close()
        return jsonify({"error": "Not found"}), 404
    course.name = data.get('name', course.name)
    course.description = data.get('description', course.description)
    course.teacher = data.get('teacher', course.teacher)
    db.commit()
    res = course.to_dict()
    db.close()
    return jsonify(res)


@api.route('/courses/<int:course_id>', methods=['DELETE'])
def delete_course(course_id):
    """Удалить курс"""
    db = SessionLocal()
    course = db.query(Course).get(course_id)
    if not course:
        db.close()
        return jsonify({"error": "Not found"}), 404
    db.delete(course)
    db.commit()
    db.close()
    return jsonify({"status": "deleted"})


# ===============================
# === RECORDS (Записи) =========
# ===============================
@api.route('/records', methods=['GET'])
def list_records():
    """Список записей"""
    q = request.args.get('q', '').strip()
    course_id = request.args.get('course_id', '').strip()
    sort = request.args.get('sort', 'default')
    db = SessionLocal()
    query = db.query(Record).join(Record.student).join(Record.course)
    if q:
        query = query.filter(or_(Student.fio.ilike(f'%{q}%'), Course.name.ilike(f'%{q}%')))
    if course_id:
        try:
            cid = int(course_id)
            query = query.filter(Record.course_id == cid)
        except ValueError:
            pass
    if sort == 'asc':
        query = query.order_by(asc(Record.date))
    elif sort == 'desc':
        query = query.order_by(desc(Record.date))
    results = [r.to_dict() for r in query.all()]
    db.close()
    return jsonify(results)


@api.route('/records', methods=['POST'])
def create_record():
    """Создать запись"""
    data = request.json
    db = SessionLocal()
    try:
        dt = datetime.fromisoformat(data['date']).date()
    except Exception:
        return jsonify({"error": "Invalid date"}), 400
    rec = Record(id_student=data['id_student'], course_id=data['course_id'], date=dt, grade=data.get('grade'))
    db.add(rec)
    db.commit()
    db.refresh(rec)
    res = rec.to_dict()
    db.close()
    return jsonify(res), 201


@api.route('/records/<int:rec_id>', methods=['PUT'])
def update_record(rec_id):
    """Обновить запись"""
    data = request.json
    db = SessionLocal()
    rec = db.query(Record).get(rec_id)
    if not rec:
        db.close()
        return jsonify({"error": "Not found"}), 404
    if 'id_student' in data:
        rec.id_student = data['id_student']
    if 'course_id' in data:
        rec.course_id = data['course_id']
    if 'date' in data:
        rec.date = datetime.fromisoformat(data['date']).date()
    if 'grade' in data:
        rec.grade = data['grade']
    db.commit()
    res = rec.to_dict()
    db.close()
    return jsonify(res)


@api.route('/records/<int:rec_id>', methods=['DELETE'])
def delete_record(rec_id):
    """Удалить запись"""
    db = SessionLocal()
    rec = db.query(Record).get(rec_id)
    if not rec:
        db.close()
        return jsonify({"error": "Not found"}), 404
    db.delete(rec)
    db.commit()
    db.close()
    return jsonify({"status": "deleted"})


# ===============================
# === WORD =====================
# ===============================
@documents_bp.route("/generate-word/<int:student_id>", methods=["GET"])
def generate_word(student_id):
    """Генерация Word-документа"""
    db = SessionLocal()
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        db.close()
        return jsonify({"error": "Студент не найден"}), 404

    doc = Document("templates/blank_zayavlenie.docx")
    for p in doc.paragraphs:
        p.text = p.text.replace("{FIO}", student.fio).replace("{PHONE}", student.phone or "")

    buffer = BytesIO()
    doc.save(buffer)
    buffer.seek(0)
    db.close()
    return send_file(buffer, as_attachment=True, download_name=f"Заявление_{student.fio}.docx")


# ===============================
# === EXCEL ====================
# ===============================
@excel_bp.route("/generate-excel", methods=["GET"])
def generate_excel():
    session = SessionLocal()

    # Загружаем записи вместе со студентом и курсом (чтобы не было DetachedInstanceError)
    records = (
        session.query(Record)
        .options(joinedload(Record.student), joinedload(Record.course))
        .all()
    )

    session.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Журнал успеваемости"

    # --- Заголовок ---
    ws.merge_cells('A1:G1')
    ws['A1'] = "ЖУРНАЛ УСПЕВАЕМОСТИ"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # --- Шапка таблицы ---
    headers = ["ID", "ФИО студента", "Дата рождения", "Телефон", "Курс", "Оценка", "Дата"]
    ws.append(headers)

    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="000000")

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # --- Данные ---
    for record in records:
        ws.append([
            record.id,
            record.student.fio if record.student else "",
            record.student.date_of_birth.strftime("%d.%m.%Y") if record.student and record.student.date_of_birth else "",
            record.student.phone if record.student else "",
            record.course.name if record.course else "",
            record.grade or "Не оценено",
            record.date.strftime("%d.%m.%Y") if record.date else ""
        ])

    # --- Автоширина ---
    for i, column_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i)
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(15, max_length + 2)

    # --- Подпись и дата ---
    ws.append([])
    ws.append([])
    ws.append(["Подпись преподавателя:", "", "", "", "Дата:", datetime.now().strftime("%d.%m.%Y")])
    ws["A" + str(ws.max_row)].font = Font(bold=True)
    ws["E" + str(ws.max_row)].font = Font(bold=True)

    # --- Сохранение ---
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f"journal_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# ===============================
# === PDF ======================
# ===============================
@pdf_bp.route("/generate-pdf/<int:student_id>", methods=["GET"])
def generate_pdf(student_id):
    db = SessionLocal()
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        db.close()
        return jsonify({"error": "Студент не найден"}), 404

    buffer = BytesIO()
    width, height = A4

    # Регистрируем шрифт
    font_path = os.path.join(current_app.root_path, "templates", "DejaVuLGCSans.ttf")
    if not os.path.exists(font_path):
        db.close()
        return jsonify({"error": f"Шрифт не найден: {font_path}"}), 500
    pdfmetrics.registerFont(TTFont('DejaVu', font_path))

    c = canvas.Canvas(buffer, pagesize=A4)

    # Логотип (прозрачный PNG)
    logo_path = os.path.join(current_app.root_path, "static", "images", "logo.png")
    if os.path.exists(logo_path):
        c.drawImage(logo_path, width - 4*cm, height - 4*cm, width=3*cm, height=3*cm, mask='auto')

        # Заголовок
    c.setFont("DejaVu", 16)
    title_y = height - 5*cm
    c.drawCentredString(width/2, title_y, "СОГЛАСИЕ НА ОБРАБОТКУ ПЕРСОНАЛЬНЫХ ДАННЫХ")


    style = ParagraphStyle(
        name="Justify",
        fontName="DejaVu",
        fontSize=12,
        leading=15,        # межстрочный интервал
        alignment=TA_JUSTIFY,
        leftIndent=2*cm,
        rightIndent=2*cm
    )

    text = f"Я, {student.fio}, даю согласие на обработку моих персональных данных в соответствии с действующим законодательством Российской Федерации."
    p = Paragraph(text, style)
    w, h = p.wrap(width - 3*cm, height)
    text_y = title_y - 1*cm  # отступ от заголовка
    p.drawOn(c, 0.5*cm, text_y - h)

    # Контактный телефон
    phone_text = f"Контактный телефон: {student.phone or 'не указан'}"
    p2 = Paragraph(phone_text, style)
    w2, h2 = p2.wrap(width - 4*cm, height)
    p2.drawOn(c, 0.5*cm, text_y - h - 0.5*cm - h2)

    # Подпись-загогулина
    signature_path = os.path.join(current_app.root_path, "static", "images", "signature.png")
    if os.path.exists(signature_path):
        c.drawImage(signature_path, 2*cm, 4*cm, width=6*cm, height=2*cm, mask='auto')
    else:
        c.drawString(2*cm, 5*cm, "Подпись: ____________________________")

    # Дата
    date_text = f"Дата: {datetime.now().strftime('%d.%m.%Y')}"
    c.drawString(2*cm, 3*cm, date_text)

    c.showPage()
    c.save()
    buffer.seek(0)
    db.close()

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"Согласие_{student.fio}.pdf",
        mimetype='application/pdf'
    )